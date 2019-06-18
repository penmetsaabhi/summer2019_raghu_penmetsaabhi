import os
import django
import openpyxl
os.environ.setdefault('DJANGO_SETTINGS_MODULE','cricket.settings')
django.setup()
from ipl.models import *
def to_matches():
    matches = openpyxl.load_workbook('matches.xlsx')
    sheet=matches.active
    for i in range(2,(sheet.max_row)+1):
        col=[]
        for j in range(1,(sheet.max_column+1)):
            col.append(sheet.cell(row=i,column=j).value)
        if(col[0]!=None):
            Matches(matchid=col[0],season=col[1],city=col[2],date=col[3],team1=col[4],team2=col[5],toss_winner=col[6],toss_decision=col[7]
                ,result=col[8],dl_applied=col[9],winner=col[10],win_by_runs=col[11],win_by_wickets=col[12],player_of_match=col[13],
                venue=col[14],umpire1=col[15],umpire2=col[16],umpire3=col[17]
                ).save()
def to_deliviers():
    matches = openpyxl.load_workbook('deliveries.xlsx')
    sheet = matches.active
    for i in range(10980, (sheet.max_row) + 1):
        col=[]
        for j in range(1, (sheet.max_column + 1)):
            col.append(sheet.cell(row=i,column=j).value)
        print(col)
        match=Matches.objects.get(matchid=col[0])
        Deliviers(match=match,match1_id=col[0], inning=col[1], batting_team=col[2], bowling_team=col[3], over=col[4],
                      ball=col[5], batsman=col[6],
                      non_striker=col[7]
                      , bowler=col[8], is_super_over=col[9], wide_runs=col[10], bye_runs=col[11], legbye_runs=col[12],
                      noball_runs=col[13],
                      penalty_runs=col[14], batsman_runs=col[15], extra_runs=col[16], total_runs=col[17],
                      player_dismissed=col[18],
                      dismissal_kind=col[19], fielder=col[20]).save()

if __name__ == '__main__':
    #to_matches()
    to_deliviers()