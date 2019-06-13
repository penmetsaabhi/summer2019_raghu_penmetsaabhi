import os
import django
import openpyxl
os.environ.setdefault('DJANGO_SETTINGS_MODULE','project1.settings')
django.setup()
from onlineapp.models import *
def toDjangoColleages():
    excel=openpyxl.load_workbook('students.xlsx')
    sheet=excel['Colleges']
    for i in range(2,(sheet.max_row)+1):
        College(name=sheet.cell(row=i,column=1).value,location=sheet.cell(row=i,column=3).value,acronym=sheet.cell(row=i,column=2).value,contact=sheet.cell(row=i,column=4).value).save()

def toDjangoStudent():
    excel = openpyxl.load_workbook('students.xlsx')
    sheet = excel['Deletions']
    """for i in range(2, (sheet.max_row) + 1):
        q=College.objects.get(acronym=(sheet.cell(row=  i, column=2)).value)
        Student(name=sheet.cell(row=i, column=1).value, email=sheet.cell(row=i, column=3).value,
                college_id=q.id, db_folder=(sheet.cell(row=i, column=4).value).lower()).save()"""
    for i in range(2, (sheet.max_row) + 1):
        q = College.objects.get(acronym=(sheet.cell(row=i, column=2)).value)
        Student(name=sheet.cell(row=i, column=1).value, email=sheet.cell(row=i, column=3).value,
                college_id=q.id, db_folder=(sheet.cell(row=i, column=4).value).lower(),dropped_out=1).save()
def marks():
    excel = openpyxl.load_workbook('Marks.xlsx')
    sheet=excel['Marks']
    for i in range(2, (sheet.max_row) + 1):
        p=(sheet.cell(row=i, column=1)).value
        p=p.split('_')[2]
        q=Student.objects.get(db_folder=p)
        MockTest1(student=q, problem1=sheet.cell(row=i, column=2).value,
                problem2=sheet.cell(row=i, column=3).value, problem3=sheet.cell(row=i, column=4).value,problem4=sheet.cell(row=i, column=5).value,total=sheet.cell(row=i, column=6).value).save()
def debugging():
    q = College.objects.all()
    print(q)

if __name__ == '__main__':
   #toDjangoColleages()
   #toDjangoStudent()
   debugging()
   # marks()